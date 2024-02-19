import pandas as pd
import openpyxl
import os
from time import time
from random import randint
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import showinfo

def Procesar_SIFERE_Excel_to_XML():

    # Timestamp
    start = int(round(time(), 0))
    #print(f"Start: {start}")

    # Nombre del archivo de excel
    #Excel = "Base.xlsx"
    Excel = askopenfilename(title="Seleccionar el archivo de excel a procesar")

    # Leer datos de cabecera
    Cuit = openpyxl.load_workbook(Excel).active["B1"].value
    Periodo = openpyxl.load_workbook(Excel).active["B2"].value
    Coef_Distribucion = openpyxl.load_workbook(Excel).active["B3"].value
    Art14 = openpyxl.load_workbook(Excel).active["B4"].value

    Gravado = openpyxl.load_workbook(Excel).active["E1"].value
    NoGravado = openpyxl.load_workbook(Excel).active["E2"].value
    Exento = openpyxl.load_workbook(Excel).active["E3"].value

    # Leer datos de la hoja de excel
    df =pd.read_excel(Excel, skiprows=5)

    # Obtener los datos de la columna jurisdicciones
    Jurisdicciones = df["Jurisdicción"].unique()

    # Crear una lista con las columnas que se van a exportar
    Columnas = ["regimenArticulo","cuacm","tratamientoFiscal","baseImponible","ajusteBaseImponible","alicuota"]

    # Crear carpeta de Resultados
    os.makedirs(f"Resultados/{Cuit}", exist_ok=True)

    for i in Jurisdicciones:
        # Filtrar los datos por jurisdicción
        df_jurisdiccion = df[df["Jurisdicción"] == i]
        
        # Eliminar la columna jurisdicciones
        del df_jurisdiccion["Jurisdicción"]
        
        # Crear el archivo xml
        df_jurisdiccion.to_xml(f"Resultados/{Cuit}/Resultado - {Cuit} - {i}.xml", index=False , attr_cols=Columnas , row_name="actividad")
            
        Cabecera = f'<ddjjSifereWeb> <cabecera id="{start}" cuit="{Cuit}" periodo="{Periodo}" timestamp="{start}" coeficienteDistribucion="{Coef_Distribucion}" articulo14="{Art14}" />'
        facturacion = f'<facturacion ingresosGravados="{Gravado}" ingresosNoGravados="{NoGravado}" ingresosExentos="{Exento}" />'
        
        # reemplazar la key "data" por "actividades"
        with open(f"Resultados/{Cuit}/Resultado - {Cuit} - {i}.xml", "r") as file:
            data = file.read()
            data = data.replace("<data>", "<actividades>")
            data = data.replace("</data>", "</actividades>")
            
            # Agregar la cabecera en la segunda linea y la facturacion en la tercera
            data = data.split("\n")
            data.insert(1, Cabecera)
            data.insert(2, facturacion)
            
            # Agregar "</ddjjSifereWeb>" al final del archivo
            data.append("</ddjjSifereWeb>")
            
            # Unir las lineas
            data = "\n".join(data)
            
            # Guardar el archivo
            with open(f"Resultados/{Cuit}/Resultado - {Cuit} - {i}.xml", "w") as file:
                file.write(data)

    showinfo("Proceso terminado", "El proceso ha terminado con éxito")